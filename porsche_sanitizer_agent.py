#!/usr/bin/env python3
"""
Porsche Sales Sanitization Agent
=================================
Le o arquivo bruto de vendas Porsche e aplica as regras descritas em
schema.md, gerando um novo arquivo .xlsx com as colunas saneadas
inseridas imediatamente apos cada coluna de origem.

Uso:
    python porsche_sanitizer_agent.py <entrada.xlsx> <saida.xlsx>
"""

import sys
import re
import datetime
import pandas as pd
import openpyxl
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# Colunas canonicas de entrada (schema.md)
# ---------------------------------------------------------------------------
INPUT_COLUMNS = [
    "sale_id", "sale_date", "customer_name", "porsche_model", "model_year",
    "sale_price", "vehicle_mileage", "payment_method", "city", "state",
    "salesperson", "delivery_status",
]

# ---------------------------------------------------------------------------
# Helpers numericos / texto por extenso
# ---------------------------------------------------------------------------
UNITS = {
    "zero": 0, "one": 1, "two": 2, "three": 3, "four": 4, "five": 5, "six": 6,
    "seven": 7, "eight": 8, "nine": 9, "ten": 10, "eleven": 11, "twelve": 12,
    "thirteen": 13, "fourteen": 14, "fifteen": 15, "sixteen": 16,
    "seventeen": 17, "eighteen": 18, "nineteen": 19,
}
TENS = {
    "twenty": 20, "thirty": 30, "forty": 40, "fifty": 50, "sixty": 60,
    "seventy": 70, "eighty": 80, "ninety": 90,
}
SCALES = {"hundred": 100, "thousand": 1000, "million": 1_000_000}


def word2num_small(phrase):
    """Soma palavras tipo 'twenty four' -> 24 (sem 'hundred'/'thousand')."""
    total = 0
    for w in phrase.split():
        if w in TENS:
            total += TENS[w]
        elif w in UNITS:
            total += UNITS[w]
        else:
            raise ValueError(f"palavra numerica desconhecida: {w}")
    return total


def text2int(phrase):
    """Algoritmo classico: 'two hundred thousand' -> 200000."""
    words = phrase.replace("-", " ").lower().split()
    current = 0
    result = 0
    found = False
    for w in words:
        if w in UNITS:
            current += UNITS[w]
            found = True
        elif w in TENS:
            current += TENS[w]
            found = True
        elif w in SCALES:
            found = True
            if w == "hundred":
                current = (current or 1) * SCALES[w]
            else:
                result += (current or 1) * SCALES[w]
                current = 0
        else:
            raise ValueError(f"palavra numerica desconhecida: {w}")
    if not found:
        raise ValueError("nenhuma palavra numerica encontrada")
    return result + current


def clean_thousand_decimal(s):
    """Resolve separadores de milhar/decimal ambiguos (US vs europeu)."""
    has_comma = "," in s
    has_dot = "." in s
    if has_comma and has_dot:
        if s.rfind(",") > s.rfind("."):
            s = s.replace(".", "").replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_comma and not has_dot:
        last = s.split(",")[-1]
        if len(last) == 2 and s.count(",") == 1:
            s = s.replace(",", ".")
        else:
            s = s.replace(",", "")
    elif has_dot and not has_comma:
        last = s.split(".")[-1]
        if len(last) == 3:
            s = s.replace(".", "")
        elif len(last) not in (1, 2):
            s = s.replace(".", "")
    return float(s)


# ---------------------------------------------------------------------------
# 1) sale_date -> SaleDateSanitized
# ---------------------------------------------------------------------------
MONTHS = {
    "january": 1, "jan": 1, "february": 2, "feb": 2, "march": 3, "mar": 3,
    "april": 4, "apr": 4, "may": 5, "june": 6, "jun": 6, "july": 7, "jul": 7,
    "august": 8, "aug": 8, "september": 9, "sep": 9, "sept": 9,
    "october": 10, "oct": 10, "november": 11, "nov": 11, "december": 12,
    "dec": 12,
}

ISO_RE = re.compile(r"^(\d{4})[-/.](\d{1,2})[-/.](\d{1,2})$")
US_RE = re.compile(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{2,4})$")
MONTHNAME_RE = re.compile(
    r"^([A-Za-z]+)\.?\s+(\d{1,2})(?:st|nd|rd|th)?,?\s+(\d{4})$"
)


def sanitize_date(raw):
    if isinstance(raw, (datetime.datetime, datetime.date)):
        try:
            return raw.strftime("%Y-%m-%d")
        except Exception:
            return "INVALID"

    if pd.isna(raw):
        return "INVALID"

    s = str(raw).strip()

    m = ISO_RE.match(s)
    if m:
        y, mo, d = (int(x) for x in m.groups())
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return "INVALID"

    m = US_RE.match(s)
    if m:
        mo, d, y = (int(x) for x in m.groups())
        if y < 100:
            y += 2000 if y < 69 else 1900
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return "INVALID"

    m = MONTHNAME_RE.match(s)
    if m:
        month_name, d, y = m.group(1).lower(), int(m.group(2)), int(m.group(3))
        mo = MONTHS.get(month_name)
        if mo is None:
            return "INVALID"
        try:
            return datetime.date(y, mo, d).isoformat()
        except ValueError:
            return "INVALID"

    return "INVALID"


# ---------------------------------------------------------------------------
# 2) porsche_model -> PorscheModelSanitized
# ---------------------------------------------------------------------------
CANONICAL_MODELS = [
    "911 Carrera", "911 Carrera S", "911 Carrera GTS", "911 Turbo",
    "911 Turbo S", "911 GT3", "911 GT3 RS", "911 Dakar", "911 Targa 4",
    "911 Targa 4S", "718 Cayman", "718 Cayman S", "718 Cayman GT4 RS",
    "718 Boxster", "718 Boxster GTS", "718 Spyder RS", "Cayenne",
    "Cayenne S", "Cayenne Coupe", "Cayenne E-Hybrid", "Cayenne Turbo",
    "Cayenne Turbo GT", "Macan", "Macan S", "Macan T", "Macan GTS",
    "Macan Electric", "Panamera", "Panamera 4", "Panamera 4S",
    "Panamera Turbo", "Panamera Turbo S", "Panamera 4 E-Hybrid", "Taycan",
    "Taycan 4S", "Taycan GTS", "Taycan Turbo", "Taycan Turbo S",
    "Taycan Cross Turismo",
]
MODEL_LOOKUP = {re.sub(r"\s+", " ", m.strip().lower()): m for m in CANONICAL_MODELS}


def smart_title(s):
    words = s.strip().split()
    out = []
    for w in words:
        if w.isupper() and len(w) <= 4:
            out.append(w)
        else:
            out.append(w[:1].upper() + w[1:].lower())
    return " ".join(out)


def sanitize_model(raw):
    if pd.isna(raw):
        return "INVALID"
    s = str(raw).strip()
    key = re.sub(r"\s+", " ", s.lower())
    if key in MODEL_LOOKUP:
        return MODEL_LOOKUP[key]
    return smart_title(s)


# ---------------------------------------------------------------------------
# 3) model_year -> ModelYearSanitized
# ---------------------------------------------------------------------------
YEAR_PAIR_RE = re.compile(r"^(\d{2})[\s-](\d{2})$")


def sanitize_year(raw):
    if pd.isna(raw):
        return "INVALID"

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        year = int(raw)
        return year if 1990 <= year <= 2035 else "INVALID"

    s = str(raw).strip()

    if s.isdigit() and len(s) == 4:
        year = int(s)
        return year if 1990 <= year <= 2035 else "INVALID"

    m = YEAR_PAIR_RE.match(s)
    if m:
        year = int(m.group(1) + m.group(2))
        return year if 1990 <= year <= 2035 else "INVALID"

    # texto por extenso
    low = s.lower().replace("-", " ")
    words = low.split()
    try:
        if "thousand" in words:
            idx = words.index("thousand")
            part1 = words[:idx]
            part2 = words[idx + 1:]
            n1 = word2num_small(" ".join(part1)) if part1 else 1
            n2 = word2num_small(" ".join(part2)) if part2 else 0
            year = n1 * 1000 + n2
        else:
            century = word2num_small(words[0])
            rest = word2num_small(" ".join(words[1:])) if len(words) > 1 else 0
            year = int(f"{century:02d}{rest:02d}")
        return year if 1990 <= year <= 2035 else "INVALID"
    except (ValueError, IndexError):
        return "INVALID"


# ---------------------------------------------------------------------------
# 4) sale_price -> SalesPriceSanitized
# ---------------------------------------------------------------------------
def sanitize_price(raw):
    if pd.isna(raw):
        return "INVALID"

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return f"{float(raw):.2f}"

    s = str(raw).strip()
    low = s.lower()
    low = re.sub(r"\busd\b", "", low)
    low = re.sub(r"\bdollars?\b", "", low)
    low = low.replace("$", "").strip()

    if not any(ch.isdigit() for ch in low):
        try:
            value = float(text2int(low))
            return f"{value:.2f}"
        except ValueError:
            return "INVALID"

    is_k = bool(re.search(r"\dk\b", low)) or low.endswith("k")
    low = re.sub(r"k\b", "", low).strip()

    try:
        value = clean_thousand_decimal(low)
    except ValueError:
        return "INVALID"

    if is_k:
        value *= 1000

    return f"{value:.2f}"


# ---------------------------------------------------------------------------
# 5) vehicle_mileage -> VehicleMileageSanitized
# ---------------------------------------------------------------------------
KM_RE = re.compile(r"\bkm\b", re.IGNORECASE)
LABEL_RE = re.compile(r"[A-Za-z:]+")


def sanitize_mileage(raw):
    if pd.isna(raw):
        return "INVALID"

    if isinstance(raw, (int, float)) and not isinstance(raw, bool):
        return int(round(raw))

    s = str(raw).strip()
    low = s.lower()

    if not any(ch.isdigit() for ch in s):
        # possivel numero por extenso, ou "new"/"zero"/etc.
        phrase = re.sub(r"\bmiles?\b|\bmi\b", "", low).strip()
        try:
            return int(text2int(phrase))
        except ValueError:
            return 0  # 'new', 'new car', valores sem numero -> 0

    is_km = bool(KM_RE.search(low))
    cleaned = LABEL_RE.sub("", s).strip()
    cleaned = cleaned.rstrip(".").strip()

    if cleaned == "":
        return 0

    try:
        value = clean_thousand_decimal(cleaned)
    except ValueError:
        return "INVALID"

    if is_km:
        value *= 0.621371

    return int(round(value))


# ---------------------------------------------------------------------------
# 6) payment_method -> PayMethodSanitized
# ---------------------------------------------------------------------------
PAYMENT_RULES = [
    ("ach", "ACH Payment"),
    ("wire", "Wire Transfer"),
    ("crypto", "Crypto Payment"),
    ("leas", "Lease"),
    ("financ", "Financing"),
    ("debit", "Debit Card"),
    ("credit", "Credit Card"),
    ("cash", "Cash"),
    ("bank", "Bank Transfer"),
]


def sanitize_payment(raw):
    if pd.isna(raw):
        return "INVALID"
    low = str(raw).strip().lower()
    for key, label in PAYMENT_RULES:
        if key in low:
            return label
    return smart_title(str(raw).strip())


# ---------------------------------------------------------------------------
# 7) city -> CitySanitized
# ---------------------------------------------------------------------------
def sanitize_city(raw):
    if pd.isna(raw):
        return "INVALID"
    return str(raw).strip().title()


# ---------------------------------------------------------------------------
# 8) state -> StateSanitized
# ---------------------------------------------------------------------------
STATE_NAME_TO_ABBR = {
    "alabama": "AL", "alaska": "AK", "arizona": "AZ", "arkansas": "AR",
    "california": "CA", "colorado": "CO", "connecticut": "CT",
    "delaware": "DE", "florida": "FL", "georgia": "GA", "hawaii": "HI",
    "idaho": "ID", "illinois": "IL", "indiana": "IN", "iowa": "IA",
    "kansas": "KS", "kentucky": "KY", "louisiana": "LA", "maine": "ME",
    "maryland": "MD", "massachusetts": "MA", "michigan": "MI",
    "minnesota": "MN", "mississippi": "MS", "missouri": "MO",
    "montana": "MT", "nebraska": "NE", "nevada": "NV",
    "new hampshire": "NH", "new jersey": "NJ", "new mexico": "NM",
    "new york": "NY", "north carolina": "NC", "north dakota": "ND",
    "ohio": "OH", "oklahoma": "OK", "oregon": "OR", "pennsylvania": "PA",
    "rhode island": "RI", "south carolina": "SC", "south dakota": "SD",
    "tennessee": "TN", "texas": "TX", "utah": "UT", "vermont": "VT",
    "virginia": "VA", "washington": "WA", "west virginia": "WV",
    "wisconsin": "WI", "wyoming": "WY",
    "district of columbia": "DC",
}
VALID_ABBR = set(STATE_NAME_TO_ABBR.values()) | {"DC"}


def sanitize_state(raw):
    if pd.isna(raw):
        return "INVALID"
    s = str(raw).strip()
    if len(s) == 2 and s.upper() in VALID_ABBR:
        return s.upper()
    name = re.sub(r"\s+", " ", s.lower())
    return STATE_NAME_TO_ABBR.get(name, "INVALID")


# ---------------------------------------------------------------------------
# 9) delivery_status -> DeliveryStatusSanitized
# ---------------------------------------------------------------------------
DELIVERY_RULES = [
    ("awaiting delivery", "Awaiting Delivery"),
    ("awaiting pickup", "Awaiting Pickup"),
    ("awaiting review", "Awaiting Review"),
    ("pending approval", "Pending Approval"),
    ("pending review", "Pending Review"),
    ("in transit", "In Transit"),
    ("cancelled", "Cancelled"),
    ("canceled", "Cancelled"),
    ("shipped", "Shipped"),
    ("deliverd", "Delivered"),
    ("delivered", "Delivered"),
    ("pending", "Pending"),
]


def sanitize_delivery(raw):
    if pd.isna(raw):
        return "INVALID"
    s = str(raw).strip().lower().replace("-", " ")
    clean = re.sub(r"[^a-z ]", "", s)
    clean = re.sub(r"\s+", " ", clean).strip()
    for key, label in DELIVERY_RULES:
        if key in clean:
            return label
    return smart_title(str(raw).strip())


# ---------------------------------------------------------------------------
# Pipeline principal
# ---------------------------------------------------------------------------
def sanitize_dataframe(df):
    df = df[INPUT_COLUMNS].copy()

    df["SaleDateSanitized"] = df["sale_date"].apply(sanitize_date)
    df["PorscheModelSanitized"] = df["porsche_model"].apply(sanitize_model)
    df["ModelYearSanitized"] = df["model_year"].apply(sanitize_year)
    df["SalesPriceSanitized"] = df["sale_price"].apply(sanitize_price)
    df["VehicleMileageSanitized"] = df["vehicle_mileage"].apply(sanitize_mileage)
    df["PayMethodSanitized"] = df["payment_method"].apply(sanitize_payment)
    df["CitySanitized"] = df["city"].apply(sanitize_city)
    df["StateSanitized"] = df["state"].apply(sanitize_state)
    df["DeliveryStatusSanitized"] = df["delivery_status"].apply(sanitize_delivery)

    ordered_cols = [
        "sale_id", "sale_date", "SaleDateSanitized",
        "customer_name",
        "porsche_model", "PorscheModelSanitized",
        "model_year", "ModelYearSanitized",
        "sale_price", "SalesPriceSanitized",
        "vehicle_mileage", "VehicleMileageSanitized",
        "payment_method", "PayMethodSanitized",
        "city", "CitySanitized",
        "state", "StateSanitized",
        "salesperson",
        "delivery_status", "DeliveryStatusSanitized",
    ]
    return df[ordered_cols]


def write_output(df, out_path):
    df.to_excel(out_path, index=False, sheet_name="Sanitized")

    wb = openpyxl.load_workbook(out_path)
    ws = wb["Sanitized"]

    header_font = Font(name="Arial", bold=True)
    body_font = Font(name="Arial")

    sanitized_cols = {
        "SaleDateSanitized", "PorscheModelSanitized", "ModelYearSanitized",
        "SalesPriceSanitized", "VehicleMileageSanitized", "PayMethodSanitized",
        "CitySanitized", "StateSanitized", "DeliveryStatusSanitized",
    }
    headers = [c.value for c in ws[1]]

    for col_idx, header in enumerate(headers, start=1):
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}1"].font = header_font
        max_len = len(str(header))
        for row in range(2, ws.max_row + 1):
            cell = ws[f"{col_letter}{row}"]
            cell.font = body_font
            max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = min(max_len + 3, 40)
        if header in sanitized_cols:
            fill = openpyxl.styles.PatternFill(
                start_color="E8F0FE", end_color="E8F0FE", fill_type="solid"
            )
            for row in range(1, ws.max_row + 1):
                ws[f"{col_letter}{row}"].fill = fill

    ws.freeze_panes = "A2"
    wb.save(out_path)


def main():
    if len(sys.argv) != 3:
        print("Uso: python porsche_sanitizer_agent.py <entrada.xlsx> <saida.xlsx>")
        sys.exit(1)

    in_path, out_path = sys.argv[1], sys.argv[2]
    df = pd.read_excel(in_path)
    sanitized = sanitize_dataframe(df)
    write_output(sanitized, out_path)

    n_invalid = (sanitized.astype(str) == "INVALID").sum().sum()
    print(f"OK: {len(sanitized)} linhas processadas -> {out_path}")
    print(f"Total de campos marcados como INVALID: {n_invalid}")


if __name__ == "__main__":
    main()
